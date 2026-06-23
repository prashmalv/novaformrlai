"""
NovoForm — Admin Panel Dialog
Accessible only to users with role='admin'.
Tabs: Team Members | Audit Log
"""
from pathlib import Path

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTabWidget, QWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QPushButton,
    QLabel, QLineEdit, QComboBox, QFormLayout, QGroupBox,
    QMessageBox, QFrame, QSizePolicy, QFileDialog, QTextEdit
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor

from src.auth.auth_manager import (
    get_all_users, add_user, deactivate_user, reactivate_user,
    reset_password, get_audit_logs, log_action,
    export_logs_csv, get_daily_log_files,
    get_db_location, set_central_db_path,
)

_NOVA_BLUE   = "#1a3a5c"
_NOVA_ACCENT = "#2c5f8a"
_WHITE       = "#ffffff"
_LIGHT       = "#dce8f5"
_HINT        = "#6c757d"
_RED         = "#c0392b"
_GREEN       = "#1e7e34"
_ROW_ALT     = "#f4f7fb"

_HDR_STYLE = f"""
    QHeaderView::section {{
        background: {_NOVA_BLUE};
        color: white;
        font-weight: 700;
        padding: 6px 4px;
        border: none;
        border-right: 1px solid #2c5f8a;
    }}
"""

_BTN = f"""
    QPushButton {{
        background:{_NOVA_BLUE}; color:{_WHITE};
        border-radius:5px; padding:6px 16px;
        font-weight:600; border:none;
    }}
    QPushButton:hover{{ background:{_NOVA_ACCENT}; }}
    QPushButton:disabled{{ background:#aab; color:#ddd; }}
"""
_BTN_RED = f"""
    QPushButton {{
        background:{_RED}; color:{_WHITE};
        border-radius:5px; padding:6px 16px;
        font-weight:600; border:none;
    }}
    QPushButton:hover{{ background:#a93226; }}
    QPushButton:disabled{{ background:#ddd; color:#aaa; }}
"""
_BTN_GREEN = f"""
    QPushButton {{
        background:{_GREEN}; color:{_WHITE};
        border-radius:5px; padding:6px 16px;
        font-weight:600; border:none;
    }}
    QPushButton:hover{{ background:#166e2c; }}
"""


class AdminPanelDialog(QDialog):

    def __init__(self, current_user: dict, parent=None):
        super().__init__(parent)
        self._user = current_user
        self.setWindowTitle("Admin Panel — NovoForm")
        self.resize(900, 580)
        self._build_ui()

    # ── Top-level layout ──────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Header bar
        hdr = QFrame()
        hdr.setFixedHeight(52)
        hdr.setStyleSheet(f"""
            background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 {_NOVA_BLUE}, stop:1 {_NOVA_ACCENT});
        """)
        hl = QHBoxLayout(hdr)
        hl.setContentsMargins(20, 0, 20, 0)
        t = QLabel("Admin Panel")
        t.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        t.setStyleSheet(f"color:{_WHITE}; background:transparent;")
        hl.addWidget(t)
        hl.addStretch()
        sub = QLabel(f"Logged in as: {self._user['full_name']} (admin)")
        sub.setStyleSheet(f"color:rgba(255,255,255,0.75); font-size:9pt; background:transparent;")
        hl.addWidget(sub)
        root.addWidget(hdr)

        # Tabs
        tabs = QTabWidget()
        tabs.setStyleSheet(f"""
            QTabBar::tab {{
                padding: 8px 24px; font-weight:600;
                color:{_HINT}; border-bottom: 3px solid transparent;
            }}
            QTabBar::tab:selected {{
                color:{_NOVA_BLUE}; border-bottom: 3px solid {_NOVA_BLUE};
            }}
            QTabWidget::pane {{ border:none; }}
        """)
        tabs.addTab(self._build_users_tab(),       "👥  Team Members")
        tabs.addTab(self._build_audit_tab(),       "📋  Audit Log")
        tabs.addTab(self._build_daily_logs_tab(),  "📁  Daily Log Files")
        tabs.addTab(self._build_network_tab(),     "🔗  Database Settings")
        root.addWidget(tabs)

    # ── Tab 1: Team Members ───────────────────────────────────────────────────

    def _build_users_tab(self) -> QWidget:
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(16)

        # Left: user table
        left = QVBoxLayout()

        refresh_btn = QPushButton("↻  Refresh")
        refresh_btn.setStyleSheet(_BTN)
        refresh_btn.setFixedWidth(110)
        refresh_btn.clicked.connect(self._load_users)
        left.addWidget(refresh_btn, alignment=Qt.AlignmentFlag.AlignRight)
        left.addSpacing(4)

        self._user_table = QTableWidget()
        self._user_table.setColumnCount(6)
        self._user_table.setHorizontalHeaderLabels(
            ["Username", "Full Name", "Role", "Status", "Created By", "Created At"])
        self._user_table.horizontalHeader().setStyleSheet(_HDR_STYLE)
        self._user_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self._user_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self._user_table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers)
        self._user_table.setAlternatingRowColors(True)
        self._user_table.setStyleSheet(
            f"alternate-background-color:{_ROW_ALT}; gridline-color:#e0e8f0;")
        self._user_table.verticalHeader().setVisible(False)
        left.addWidget(self._user_table)

        # Action buttons for selected user
        act_row = QHBoxLayout()
        self._deact_btn = QPushButton("Deactivate User")
        self._deact_btn.setStyleSheet(_BTN_RED)
        self._deact_btn.clicked.connect(self._deactivate_selected)
        self._react_btn = QPushButton("Reactivate User")
        self._react_btn.setStyleSheet(_BTN_GREEN)
        self._react_btn.clicked.connect(self._reactivate_selected)
        self._rpw_btn = QPushButton("Reset Password")
        self._rpw_btn.setStyleSheet(_BTN)
        self._rpw_btn.clicked.connect(self._reset_pw_selected)
        act_row.addWidget(self._deact_btn)
        act_row.addWidget(self._react_btn)
        act_row.addWidget(self._rpw_btn)
        act_row.addStretch()
        left.addLayout(act_row)

        lay.addLayout(left, stretch=3)

        # Right: add user form
        grp = QGroupBox("Add New User")
        grp.setStyleSheet(f"""
            QGroupBox {{
                font-weight:700; color:{_NOVA_BLUE};
                border:1.5px solid {_LIGHT}; border-radius:8px;
                margin-top:8px; padding:12px 14px;
            }}
            QGroupBox::title {{ subcontrol-origin:margin; left:10px; }}
        """)
        f = QFormLayout(grp)
        f.setSpacing(10)

        self._new_uname = QLineEdit(); self._new_uname.setPlaceholderText("e.g. rajan.kumar")
        self._new_fname = QLineEdit(); self._new_fname.setPlaceholderText("e.g. Rajan Kumar")
        self._new_pw    = QLineEdit(); self._new_pw.setEchoMode(QLineEdit.EchoMode.Password)
        self._new_pw.setPlaceholderText("Min 6 characters")
        self._new_role  = QComboBox(); self._new_role.addItems(["user", "admin"])

        for field in (self._new_uname, self._new_fname, self._new_pw):
            field.setStyleSheet("""
                QLineEdit { border:1.5px solid #c8d8e8; border-radius:5px;
                            padding:4px 8px; font-size:9pt; }
                QLineEdit:focus { border-color:#2c5f8a; }
            """)

        f.addRow("Username:",   self._new_uname)
        f.addRow("Full Name:",  self._new_fname)
        f.addRow("Password:",   self._new_pw)
        f.addRow("Role:",       self._new_role)

        self._add_msg = QLabel("")
        self._add_msg.setWordWrap(True)
        self._add_msg.setStyleSheet("font-size:8pt;")
        f.addRow("", self._add_msg)

        add_btn = QPushButton("Add User")
        add_btn.setStyleSheet(_BTN)
        add_btn.clicked.connect(self._do_add_user)
        f.addRow("", add_btn)

        right_lay = QVBoxLayout()
        right_lay.addWidget(grp)
        right_lay.addStretch()
        lay.addLayout(right_lay, stretch=1)

        self._load_users()
        return w

    def _load_users(self):
        users = get_all_users()
        self._user_table.setRowCount(len(users))
        for r, u in enumerate(users):
            active = u["active"] == 1
            items = [
                u["username"], u["full_name"], u["role"],
                "Active" if active else "Inactive",
                u["created_by"], u["created_at"]
            ]
            for c, val in enumerate(items):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                if c == 3:  # Status
                    item.setForeground(QColor(_GREEN if active else _RED))
                    item.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                if c == 2 and val == "admin":
                    item.setForeground(QColor(_NOVA_BLUE))
                    item.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                self._user_table.setItem(r, c, item)

    def _selected_username(self) -> str | None:
        rows = self._user_table.selectedItems()
        if not rows:
            return None
        row = self._user_table.currentRow()
        return self._user_table.item(row, 0).text()

    def _deactivate_selected(self):
        uname = self._selected_username()
        if not uname:
            QMessageBox.information(self, "Select User", "Select a user row first.")
            return
        if uname == self._user["username"]:
            QMessageBox.warning(self, "Not Allowed", "You cannot deactivate your own account.")
            return
        ok = QMessageBox.question(
            self, "Confirm",
            f"Deactivate user '{uname}'?\nThey will not be able to log in.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if ok != QMessageBox.StandardButton.Yes:
            return
        success, msg = deactivate_user(uname)
        if success:
            log_action(self._user["username"], self._user["full_name"],
                       "USER_DEACTIVATED", f"Deactivated user: {uname}")
        QMessageBox.information(self, "Result", msg)
        self._load_users()

    def _reactivate_selected(self):
        uname = self._selected_username()
        if not uname:
            QMessageBox.information(self, "Select User", "Select a user row first.")
            return
        success, msg = reactivate_user(uname)
        if success:
            log_action(self._user["username"], self._user["full_name"],
                       "USER_REACTIVATED", f"Reactivated user: {uname}")
        QMessageBox.information(self, "Result", msg)
        self._load_users()

    def _reset_pw_selected(self):
        uname = self._selected_username()
        if not uname:
            QMessageBox.information(self, "Select User", "Select a user row first.")
            return
        from PyQt6.QtWidgets import QInputDialog
        new_pw, ok = QInputDialog.getText(
            self, "Reset Password",
            f"Enter new password for '{uname}':",
            QLineEdit.EchoMode.Password
        )
        if not ok or not new_pw.strip():
            return
        if len(new_pw.strip()) < 6:
            QMessageBox.warning(self, "Too Short", "Password must be at least 6 characters.")
            return
        success, msg = reset_password(uname, new_pw.strip(),
                                      self._user["username"])
        QMessageBox.information(self, "Result", msg)

    def _do_add_user(self):
        uname = self._new_uname.text().strip()
        fname = self._new_fname.text().strip()
        pw    = self._new_pw.text()
        role  = self._new_role.currentText()

        if len(pw) < 6:
            self._add_msg.setStyleSheet(f"color:{_RED}; font-size:8pt;")
            self._add_msg.setText("Password must be at least 6 characters.")
            return

        success, msg = add_user(uname, fname, pw, role, self._user["username"])
        if success:
            log_action(self._user["username"], self._user["full_name"],
                       "USER_ADDED",
                       f"Added user: {uname} ({fname}), role={role}")
            self._add_msg.setStyleSheet(f"color:{_GREEN}; font-size:8pt;")
            self._new_uname.clear(); self._new_fname.clear(); self._new_pw.clear()
            self._load_users()
        else:
            self._add_msg.setStyleSheet(f"color:{_RED}; font-size:8pt;")
        self._add_msg.setText(msg)

    # ── Tab 2: Audit Log ──────────────────────────────────────────────────────

    def _build_audit_tab(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(10)

        # Filter + action row
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Filter by user:"))
        self._audit_user_filter = QLineEdit()
        self._audit_user_filter.setPlaceholderText("Leave blank for all users")
        self._audit_user_filter.setFixedWidth(180)
        self._audit_user_filter.setStyleSheet(
            "QLineEdit{border:1.5px solid #c8d8e8;border-radius:5px;"
            "padding:4px 8px;font-size:9pt;}")
        filter_row.addWidget(self._audit_user_filter)
        filter_row.addSpacing(8)
        filter_row.addWidget(QLabel("Rows:"))
        self._audit_limit = QComboBox()
        self._audit_limit.addItems(["100", "250", "500", "1000"])
        self._audit_limit.setCurrentText("250")
        self._audit_limit.setFixedWidth(75)
        filter_row.addWidget(self._audit_limit)
        filter_row.addSpacing(8)

        ref_btn = QPushButton("↻  Refresh")
        ref_btn.setStyleSheet(_BTN); ref_btn.setFixedWidth(100)
        ref_btn.clicked.connect(self._load_audit)
        filter_row.addWidget(ref_btn)
        filter_row.addSpacing(6)

        exp_csv_btn = QPushButton("⬇  Export CSV")
        exp_csv_btn.setStyleSheet(_BTN); exp_csv_btn.setFixedWidth(120)
        exp_csv_btn.clicked.connect(self._export_csv)
        filter_row.addWidget(exp_csv_btn)

        exp_xl_btn = QPushButton("⬇  Export Excel")
        exp_xl_btn.setStyleSheet(_BTN); exp_xl_btn.setFixedWidth(130)
        exp_xl_btn.clicked.connect(self._export_excel)
        filter_row.addWidget(exp_xl_btn)

        filter_row.addStretch()
        lay.addLayout(filter_row)

        # 7 columns: Timestamp | User | Full Name | Action | Details | Hostname | IP
        self._audit_table = QTableWidget()
        self._audit_table.setColumnCount(7)
        self._audit_table.setHorizontalHeaderLabels(
            ["Timestamp", "User", "Full Name", "Action", "Details", "Hostname", "IP Address"])
        self._audit_table.horizontalHeader().setStyleSheet(_HDR_STYLE)
        hh = self._audit_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self._audit_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._audit_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._audit_table.setAlternatingRowColors(True)
        self._audit_table.setStyleSheet(
            f"alternate-background-color:{_ROW_ALT}; gridline-color:#e0e8f0;")
        self._audit_table.verticalHeader().setVisible(False)
        lay.addWidget(self._audit_table)

        self._audit_count_lbl = QLabel("")
        self._audit_count_lbl.setStyleSheet(f"color:{_HINT}; font-size:8pt;")
        lay.addWidget(self._audit_count_lbl)

        self._audit_rows_cache: list[dict] = []
        self._load_audit()
        return w

    # Colour per action
    _ACTION_COLORS = {
        "LOGIN":             "#1e7e34",
        "LOGOUT":            "#5a6268",
        "DXF_IMPORT":        "#0c5460",
        "PDF_IMPORT":        "#0c5460",
        "BOQ_COMPUTED":      "#004085",
        "PDF_EXPORTED":      "#533f03",
        "EXCEL_EXPORTED":    "#533f03",
        "ELEMENT_ADDED":     "#155724",
        "ELEMENT_EDITED":    "#856404",
        "ELEMENT_DELETED":   "#721c24",
        "USER_ADDED":        "#1e7e34",
        "USER_DEACTIVATED":  "#721c24",
        "USER_REACTIVATED":  "#1e7e34",
        "PASSWORD_RESET":    "#533f03",
        "ADMIN_PANEL_OPENED":"#495057",
    }

    def _load_audit(self):
        limit = int(self._audit_limit.currentText())
        ufilt = self._audit_user_filter.text().strip().lower()
        rows  = get_audit_logs(limit=limit, username_filter=ufilt or "")
        self._audit_rows_cache = rows

        self._audit_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            vals = [
                row.get("timestamp", ""),
                row.get("username", ""),
                row.get("full_name", ""),
                row.get("action", ""),
                row.get("details") or "",
                row.get("hostname") or row.get("host") or "",
                row.get("ip_address") or "",
            ]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                if c == 3:
                    color = self._ACTION_COLORS.get(str(val), _NOVA_BLUE)
                    item.setForeground(QColor(color))
                    item.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
                self._audit_table.setItem(r, c, item)

        self._audit_count_lbl.setText(
            f"Showing {len(rows)} record(s)"
            + (f" · user filter: '{ufilt}'" if ufilt else ""))

    def _export_csv(self):
        if not self._audit_rows_cache:
            QMessageBox.information(self, "No Data", "Load audit data first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Audit Log as CSV", "audit_log.csv",
            "CSV Files (*.csv)")
        if not path:
            return
        csv_text = export_logs_csv(self._audit_rows_cache)
        try:
            Path(path).write_text(csv_text, encoding="utf-8-sig")
            QMessageBox.information(self, "Exported",
                                    f"CSV saved:\n{path}")
            log_action(self._user["username"], self._user["full_name"],
                       "AUDIT_EXPORTED", f"CSV: {Path(path).name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _export_excel(self):
        if not self._audit_rows_cache:
            QMessageBox.information(self, "No Data", "Load audit data first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Audit Log as Excel", "audit_log.xlsx",
            "Excel Files (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Audit Log"

            hdr_fill = PatternFill("solid", fgColor="1A3A5C")
            hdr_font = Font(bold=True, color="FFFFFF", size=10)
            cols = ["Timestamp", "User", "Full Name", "Action",
                    "Details", "Hostname", "IP Address"]
            for c, h in enumerate(cols, 1):
                cell = ws.cell(1, c, h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")

            keys = ["timestamp", "username", "full_name", "action",
                    "details", "hostname", "ip_address"]
            for r, row in enumerate(self._audit_rows_cache, 2):
                for c, k in enumerate(keys, 1):
                    v = row.get(k) or row.get("host") if k == "hostname" else row.get(k)
                    ws.cell(r, c, str(v or ""))

            # Auto-fit column widths
            for col in ws.columns:
                max_w = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 50)

            wb.save(path)
            QMessageBox.information(self, "Exported", f"Excel saved:\n{path}")
            log_action(self._user["username"], self._user["full_name"],
                       "AUDIT_EXPORTED", f"Excel: {Path(path).name}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ── Tab 3: Daily Log Files ────────────────────────────────────────────────

    def _build_daily_logs_tab(self) -> QWidget:
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(12)

        # Left: file list
        left = QVBoxLayout()

        # Location hint
        from src.auth.auth_manager import _appdata_logs_root
        loc_lbl = QLabel(f"📁 Location: {_appdata_logs_root()}")
        loc_lbl.setStyleSheet(
            f"color:{_HINT}; font-size:8pt; font-style:italic;")
        loc_lbl.setWordWrap(True)
        left.addWidget(loc_lbl)
        left.addSpacing(4)

        # Filter by user
        frow = QHBoxLayout()
        frow.addWidget(QLabel("User:"))
        self._log_user_filter = QComboBox()
        self._log_user_filter.addItem("All users")
        for u in get_all_users():
            self._log_user_filter.addItem(u["username"])
        self._log_user_filter.currentIndexChanged.connect(self._load_log_file_list)
        frow.addWidget(self._log_user_filter)
        frow.addStretch()
        ref2 = QPushButton("↻")
        ref2.setFixedWidth(30); ref2.setStyleSheet(_BTN)
        ref2.clicked.connect(self._load_log_file_list)
        frow.addWidget(ref2)
        left.addLayout(frow)
        left.addSpacing(4)

        self._log_file_list = QTableWidget()
        self._log_file_list.setColumnCount(3)
        self._log_file_list.setHorizontalHeaderLabels(["User", "Date", "Size"])
        self._log_file_list.horizontalHeader().setStyleSheet(_HDR_STYLE)
        self._log_file_list.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self._log_file_list.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self._log_file_list.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.ResizeToContents)
        self._log_file_list.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self._log_file_list.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers)
        self._log_file_list.verticalHeader().setVisible(False)
        self._log_file_list.setAlternatingRowColors(True)
        self._log_file_list.setStyleSheet(
            f"alternate-background-color:{_ROW_ALT};")
        self._log_file_list.setFixedWidth(260)
        self._log_file_list.itemSelectionChanged.connect(self._show_log_file)
        left.addWidget(self._log_file_list)

        btn_row = QHBoxLayout()
        dl_btn = QPushButton("⬇  Save Copy")
        dl_btn.setStyleSheet(_BTN); dl_btn.setFixedWidth(120)
        dl_btn.clicked.connect(self._save_log_file_copy)
        btn_row.addWidget(dl_btn)
        btn_row.addStretch()
        left.addLayout(btn_row)

        lay.addLayout(left)

        # Right: file content viewer
        right = QVBoxLayout()
        self._log_file_title = QLabel("Select a file to view its contents")
        self._log_file_title.setStyleSheet(
            f"color:{_NOVA_BLUE}; font-weight:700; font-size:9pt;")
        right.addWidget(self._log_file_title)

        self._log_viewer = QTextEdit()
        self._log_viewer.setReadOnly(True)
        self._log_viewer.setFont(QFont("Courier New", 8))
        self._log_viewer.setStyleSheet(
            "background:#f8fafc; border:1px solid #c8d8e8; border-radius:4px;")
        right.addWidget(self._log_viewer)
        lay.addLayout(right, stretch=1)

        # (username, path) tuples
        self._daily_log_files: list[tuple[str, Path]] = []
        self._load_log_file_list()
        return w

    def _load_log_file_list(self):
        sel = self._log_user_filter.currentText()
        ufilter = "" if sel == "All users" else sel
        self._daily_log_files = get_daily_log_files(username=ufilter)
        self._log_file_list.setRowCount(len(self._daily_log_files))
        for r, (uname, p) in enumerate(self._daily_log_files):
            size_kb = p.stat().st_size / 1024
            self._log_file_list.setItem(r, 0, QTableWidgetItem(uname))
            self._log_file_list.setItem(r, 1, QTableWidgetItem(p.stem))
            self._log_file_list.setItem(
                r, 2, QTableWidgetItem(f"{size_kb:.1f} KB"))
        if self._daily_log_files:
            self._log_file_list.selectRow(0)
        else:
            self._log_viewer.setPlainText("No log files found.")

    def _show_log_file(self):
        row = self._log_file_list.currentRow()
        if row < 0 or row >= len(self._daily_log_files):
            return
        uname, p = self._daily_log_files[row]
        self._log_file_title.setText(f"📄  {uname} / {p.name}")
        try:
            self._log_viewer.setPlainText(p.read_text(encoding="utf-8"))
            self._log_viewer.moveCursor(
                self._log_viewer.textCursor().MoveOperation.End)
        except Exception as e:
            self._log_viewer.setPlainText(f"Error reading file:\n{e}")

    def _save_log_file_copy(self):
        row = self._log_file_list.currentRow()
        if row < 0 or row >= len(self._daily_log_files):
            QMessageBox.information(self, "Select File", "Select a file first.")
            return
        uname, src = self._daily_log_files[row]
        default_name = f"{uname}_{src.stem}.txt"
        dst, _ = QFileDialog.getSaveFileName(
            self, "Save Log File Copy", default_name, "Text Files (*.txt)")
        if not dst:
            return
        try:
            Path(dst).write_bytes(src.read_bytes())
            QMessageBox.information(self, "Saved", f"Log file saved to:\n{dst}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ── Tab 4: Database / Network Settings ───────────────────────────────────

    def _build_network_tab(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(24, 24, 24, 24)
        lay.setSpacing(16)

        # ── Current status ────────────────────────────────────────────────────
        status_grp = QGroupBox("Current Database Location")
        status_grp.setStyleSheet(f"""
            QGroupBox {{
                font-weight:700; color:{_NOVA_BLUE};
                border:1.5px solid {_LIGHT}; border-radius:8px;
                margin-top:8px; padding:14px 16px;
            }}
            QGroupBox::title {{ subcontrol-origin:margin; left:10px; }}
        """)
        sg = QVBoxLayout(status_grp)

        self._db_status_lbl = QLabel()
        self._db_status_lbl.setWordWrap(True)
        self._db_status_lbl.setStyleSheet("font-size:10pt;")
        sg.addWidget(self._db_status_lbl)

        self._db_path_lbl = QLabel()
        self._db_path_lbl.setWordWrap(True)
        self._db_path_lbl.setStyleSheet(
            f"font-family:Courier New; font-size:9pt; color:{_NOVA_BLUE}; "
            f"background:#f0f4f8; padding:6px; border-radius:4px;")
        sg.addWidget(self._db_path_lbl)

        self._db_reach_lbl = QLabel()
        sg.addWidget(self._db_reach_lbl)

        lay.addWidget(status_grp)

        # ── Set central path ──────────────────────────────────────────────────
        cfg_grp = QGroupBox("Set Central Database Path  (shared folder on admin's machine)")
        cfg_grp.setStyleSheet(status_grp.styleSheet())
        cg = QVBoxLayout(cfg_grp)

        how_lbl = QLabel(
            "<b>How to use:</b><br>"
            "1. On <b>admin's machine</b> — share a folder (e.g. <code>C:\\NovoFormDB\\</code>)<br>"
            "2. On <b>each employee machine</b> — enter the UNC or mapped-drive path below<br>"
            "3. Click <b>Save &amp; Apply</b> — app will use shared DB from next login<br><br>"
            "Example paths:<br>"
            "&nbsp;&nbsp;Windows UNC&nbsp;: <code>\\\\ADMIN-PC\\NovoFormDB\\novoform_auth.db</code><br>"
            "&nbsp;&nbsp;Mapped drive : <code>Z:\\novoform_auth.db</code><br>"
            "&nbsp;&nbsp;Leave blank to use local DB (admin's own machine)"
        )
        how_lbl.setWordWrap(True)
        how_lbl.setStyleSheet("font-size:9pt; color:#333; padding:4px;")
        how_lbl.setTextFormat(Qt.TextFormat.RichText)
        cg.addWidget(how_lbl)
        cg.addSpacing(8)

        path_row = QHBoxLayout()
        self._central_path_edit = QLineEdit()
        self._central_path_edit.setPlaceholderText(
            r"e.g. \\ADMIN-PC\NovoFormDB\novoform_auth.db  (blank = use local)")
        self._central_path_edit.setStyleSheet("""
            QLineEdit { border:1.5px solid #c8d8e8; border-radius:5px;
                        padding:6px 10px; font-family:Courier New; font-size:9pt; }
            QLineEdit:focus { border-color:#2c5f8a; }
        """)
        path_row.addWidget(self._central_path_edit)

        browse_btn = QPushButton("Browse…")
        browse_btn.setStyleSheet(_BTN)
        browse_btn.setFixedWidth(90)
        browse_btn.clicked.connect(self._browse_db_path)
        path_row.addWidget(browse_btn)
        cg.addLayout(path_row)
        cg.addSpacing(8)

        btn_row = QHBoxLayout()
        save_btn = QPushButton("Save & Apply")
        save_btn.setStyleSheet(_BTN)
        save_btn.setFixedWidth(130)
        save_btn.clicked.connect(self._save_db_path)
        btn_row.addWidget(save_btn)

        test_btn = QPushButton("Test Connection")
        test_btn.setStyleSheet(_BTN)
        test_btn.setFixedWidth(140)
        test_btn.clicked.connect(self._test_db_path)
        btn_row.addWidget(test_btn)

        reset_btn = QPushButton("Use Local DB")
        reset_btn.setStyleSheet(_BTN_RED)
        reset_btn.setFixedWidth(120)
        reset_btn.clicked.connect(self._reset_to_local_db)
        btn_row.addWidget(reset_btn)
        btn_row.addStretch()
        cg.addLayout(btn_row)

        self._db_msg_lbl = QLabel("")
        self._db_msg_lbl.setStyleSheet("font-size:9pt;")
        self._db_msg_lbl.setWordWrap(True)
        cg.addWidget(self._db_msg_lbl)
        lay.addWidget(cfg_grp)

        # ── Info box ──────────────────────────────────────────────────────────
        info = QLabel(
            "ℹ  <b>Admin machine setup:</b> Share a folder and place (or let NovoForm create) "
            "the DB file there. No server required — all machines write directly to the shared file.<br>"
            "ℹ  <b>Works best on a stable LAN.</b> If the shared folder is unreachable, "
            "login will fail. Use 'Use Local DB' to fall back to standalone mode."
        )
        info.setWordWrap(True)
        info.setStyleSheet(
            f"background:{_LIGHT}; color:{_NOVA_BLUE}; border-radius:6px; "
            f"padding:10px 14px; font-size:9pt;")
        info.setTextFormat(Qt.TextFormat.RichText)
        lay.addWidget(info)
        lay.addStretch()

        self._refresh_db_status()
        return w

    def _refresh_db_status(self):
        info = get_db_location()
        self._db_path_lbl.setText(info["path"])
        if info["is_central"]:
            self._db_status_lbl.setText("🔗  Mode: <b>Central shared database</b>")
            self._db_status_lbl.setStyleSheet(f"font-size:10pt; color:{_GREEN};")
        else:
            self._db_status_lbl.setText("💻  Mode: <b>Local database (this machine only)</b>")
            self._db_status_lbl.setStyleSheet(f"font-size:10pt; color:{_NOVA_BLUE};")
        if info["reachable"]:
            self._db_reach_lbl.setText("✅  Path is reachable")
            self._db_reach_lbl.setStyleSheet(f"color:{_GREEN}; font-size:9pt;")
        else:
            self._db_reach_lbl.setText("❌  Path NOT reachable — check network / folder sharing")
            self._db_reach_lbl.setStyleSheet(f"color:{_RED}; font-size:9pt;")
        self._central_path_edit.setText(
            info["path"] if info["is_central"] else "")

    def _browse_db_path(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Select or Create Database File",
            "novoform_auth.db", "SQLite Database (*.db)")
        if path:
            self._central_path_edit.setText(path)

    def _test_db_path(self):
        raw = self._central_path_edit.text().strip()
        if not raw:
            self._db_msg_lbl.setStyleSheet(f"color:{_NOVA_BLUE}; font-size:9pt;")
            self._db_msg_lbl.setText("ℹ  Empty path = local mode. No test needed.")
            return
        p = Path(raw)
        if not p.parent.exists():
            self._db_msg_lbl.setStyleSheet(f"color:{_RED}; font-size:9pt;")
            self._db_msg_lbl.setText(f"❌  Folder not found: {p.parent}")
            return
        # Try opening/creating a connection
        try:
            import sqlite3 as _sq
            con = _sq.connect(str(p))
            con.execute("SELECT 1")
            con.close()
            self._db_msg_lbl.setStyleSheet(f"color:{_GREEN}; font-size:9pt;")
            self._db_msg_lbl.setText(f"✅  Connection successful: {p}")
        except Exception as e:
            self._db_msg_lbl.setStyleSheet(f"color:{_RED}; font-size:9pt;")
            self._db_msg_lbl.setText(f"❌  Cannot open database: {e}")

    def _save_db_path(self):
        raw = self._central_path_edit.text().strip()
        ok, msg = set_central_db_path(raw)
        if ok:
            self._db_msg_lbl.setStyleSheet(f"color:{_GREEN}; font-size:9pt;")
            log_action(self._user["username"], self._user["full_name"],
                       "DB_PATH_CHANGED",
                       f"central_db_path set to: '{raw or '(local)'}'")
        else:
            self._db_msg_lbl.setStyleSheet(f"color:{_RED}; font-size:9pt;")
        self._db_msg_lbl.setText(msg)
        self._refresh_db_status()

    def _reset_to_local_db(self):
        ok = QMessageBox.question(
            self, "Revert to Local",
            "Revert to local database?\n\nThis machine will no longer use the shared DB.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ok != QMessageBox.StandardButton.Yes:
            return
        self._central_path_edit.clear()
        self._save_db_path()
