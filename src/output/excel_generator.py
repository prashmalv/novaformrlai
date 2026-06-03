"""
Excel BOQ Generator — Nova Formworks updated format (2025).
Sheets: BOQ (panel quantities) + DAYS BOQ (reuse schedule) + QUOTATION (pricing)
"""
from __future__ import annotations

import datetime
import re
from collections import defaultdict, OrderedDict
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models.element import ElementBOQ, ProjectBOQ

_LOGO_PATH = Path(__file__).parent.parent.parent / "assets" / "images" / "NovaLogo.png"

# ── Brand colors (2025) ───────────────────────────────────────────────────────
_PURPLE     = "7D2354"   # primary header
_NAVY       = "3A516F"   # secondary
_SMOKE      = "F4F4F4"   # alternating rows
_LIGHT_P    = "F5E8EF"   # very light purple (OC rows)
_NIGHT      = "232323"   # body text
_WHITE      = "FFFFFF"

_thin  = Side(style="thin",   color="CCCCCC")
_thick = Side(style="medium", color="000000")


def _border_thin() -> Border:
    return Border(top=_thin, bottom=_thin, left=_thin, right=_thin)


def _hdr(ws, row, col, value, bg=_PURPLE, fc=_WHITE, bold=True, wrap=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fc, size=8)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _border_thin()
    return c


def _cell(ws, row, col, value, bold=False, fill=None, align="center",
          fmt=None, color=_NIGHT, size=8):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=color, size=size)
    if fill:
        c.fill  = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        c.number_format = fmt
    c.border    = _border_thin()
    return c


def _fmt_panel(size_label: str) -> str:
    if size_label.startswith('OC'):
        m = re.match(r'OC(\d+)X(\d+)', size_label)
        return f"Outer Corner {m.group(1)}*{m.group(2)}" if m else size_label
    if size_label.startswith('IC'):
        m = re.match(r'IC(\d+)X(\d+)', size_label)
        return f"Inner Corner {m.group(1)}*{m.group(2)}" if m else size_label
    m = re.match(r'(\d+)X(\d+)', size_label)
    return f"Panel {m.group(1)}*{m.group(2)}" if m else size_label


def _fmt_acc(size_label: str) -> tuple[str, str]:
    sl = size_label.upper()
    if 'WALLER' in sl:
        m = re.search(r'(\d+\.?\d*)\s*M', sl)
        return f"accessories - waller - {m.group(1) if m else '?'}", "nos"
    if 'TIE ROD' in sl or 'TIEROD' in sl:
        m = re.search(r'\((\d+\.?\d*)M\)', sl) or re.search(r'(\d+\.?\d*)M', sl)
        return f"accessories - tie_rod - {m.group(1) if m else '?'} (16 MM)", "nos"
    if 'WING NUT' in sl:
        return "accessories - wing_nut", "nos"
    if 'PVC CONE' in sl or 'CONE' in sl:
        return "accessories - pvc_cone", "nos"
    if 'ANCHOR' in sl or ('NUT' in sl and 'WING' not in sl):
        return "accessories - anchor_nut - 100", "nos"
    if 'PIN' in sl:
        m = re.search(r'(\d+)', sl)
        return f"accessories - pin_{m.group(1) if m else '?'}mm", "nos"
    return f"accessories - {size_label.lower()}", "nos"


def _group_boqs(element_boqs: list) -> list[dict]:
    groups: dict[tuple, dict] = OrderedDict()
    for boq in element_boqs:
        el  = boq.element
        key = (el.element_type, round(el.length_mm), round(el.width_mm))
        if key not in groups:
            groups[key] = {'boq': boq, 'count': 0, 'labels': [], 'height_mm': el.height_mm}
        groups[key]['count']  += max(1, el.quantity)
        groups[key]['labels'].append(el.label)
    return list(groups.values())


# ══════════════════════════════════════════════════════════════════════════════
# BOQ sheet
# ══════════════════════════════════════════════════════════════════════════════

def _write_boq_sheet(wb, project: ProjectBOQ, boq_number: str = None):
    ws  = wb.create_sheet("FORMWORK BOQ")
    row = 1

    # ── Document header ────────────────────────────────────────────────────────
    date_str   = project.date or datetime.date.today().strftime("%d/%m/%Y")
    boq_num    = boq_number or f"OP-ID-{abs(hash(project.project_name)) % 10000:04d}"

    # Logo (A1:A3)
    ws.merge_cells("A1:A3")
    ws.column_dimensions["A"].width = 5
    if _LOGO_PATH.exists():
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(str(_LOGO_PATH))
            img.width, img.height = 90, 28
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            ws["A1"].value = "NOVA"

    # Title block
    ws.merge_cells("B1:I1")
    c = ws["B1"]
    c.value = "FORMWORK BOQ"
    c.font  = Font(bold=True, size=14, color=_NIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    ws.merge_cells("B2:I2")
    ws["B2"].value = f"BOQ Number: {boq_num}     Date: {date_str}"
    ws["B2"].alignment = Alignment(horizontal="center")
    ws["B2"].font = Font(size=9, color=_NIGHT)

    ws.merge_cells("B3:I3")
    ws["B3"].value = ""
    ws.row_dimensions[3].height = 6
    row = 4

    # ── From / To ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"].value = "From:"
    ws[f"B{row}"].font  = Font(bold=True, size=8)
    ws.merge_cells(f"F{row}:I{row}")
    ws[f"F{row}"].value = "To:"
    ws[f"F{row}"].font  = Font(bold=True, size=8, color=_NIGHT)
    ws[f"F{row}"].alignment = Alignment(horizontal="right")
    row += 1

    from_lines = [
        "NOVA FORMWORKS",
        "A-7/121-124 South Side of GT Road Indl. Area",
        "Ghaziabad (UP)",
        f"Email: rawatkaran7412@gmail.com",
        f"Mobile: 7819998963",
    ]
    to_lines = [
        project.client_name or "—",
        project.client_address or "—",
        "",
        f"Email: {getattr(project, 'client_email', '') or '-'}",
        f"Phone: {getattr(project, 'client_phone', '') or '-'}",
    ]
    for fl, tl in zip(from_lines, to_lines):
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].value = fl
        ws[f"B{row}"].font  = Font(bold=(row == 5), size=8)
        ws.merge_cells(f"F{row}:I{row}")
        ws[f"F{row}"].value = tl
        ws[f"F{row}"].font  = Font(bold=(row == 5), size=8)
        ws[f"F{row}"].alignment = Alignment(horizontal="right")
        row += 1

    ws.row_dimensions[row].height = 6
    row += 1

    # ── BOQ DETAILS header ────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:I{row}")
    c = ws[f"B{row}"]
    c.value = "BOQ DETAILS:"
    c.font  = Font(bold=True, size=9, color=_WHITE)
    c.fill  = PatternFill("solid", fgColor=_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1

    # ── Per-dimension element blocks ──────────────────────────────────────────
    groups = _group_boqs(project.element_boqs)
    col_hdrs = ['PRODUCT', 'Qty', 'UOM', 'No of Set', 'Total Qty',
                'Unit Area\n(SqM)', 'Total Area\n(SqM)']
    cw_map   = {'B': 28, 'C': 9, 'D': 9, 'E': 12, 'F': 12, 'G': 14, 'H': 14, 'I': 6}

    for g in groups:
        boq      = g['boq']
        no_sets  = g['count']
        el       = boq.element
        h_str    = f"{int(g['height_mm']):,}MM"
        dim_str  = f"{int(el.length_mm)}X{int(el.width_mm)}"
        req_type = el.element_type.value.lower()

        # Client Requirement row
        ws.merge_cells(f"B{row}:H{row}")
        c = ws[f"B{row}"]
        c.value = (f"Client Requirement: {req_type}  |  Dimension: {dim_str}  |  "
                   f"FormWork Area: -  |  Height: {h_str}")
        c.font  = Font(bold=True, size=8, color=_NIGHT)
        c.fill  = PatternFill("solid", fgColor=_SMOKE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

        # Column headers
        for ci, h in enumerate(col_hdrs, start=2):
            _hdr(ws, row, ci, h, bg=_PURPLE, fc=_WHITE)
        ws.row_dimensions[row].height = 28
        row += 1

        total_area = 0.0
        first      = True
        for panel in boq.panels:
            label     = _fmt_panel(panel.size_label)
            qty       = panel.quantity
            unit_a    = round(panel.width_mm * panel.height_mm / 1_000_000, 2)
            total_qty = qty * no_sets
            row_area  = round(unit_a * total_qty, 2)
            total_area += row_area
            is_corner = panel.is_corner or panel.is_inner_corner
            fill = _LIGHT_P if is_corner else (None if (row % 2 == 0) else _SMOKE)

            _cell(ws, row, 2, label,      fill=fill, align="left", bold=is_corner)
            _cell(ws, row, 3, f"{qty:.2f}", fill=fill)
            _cell(ws, row, 4, "nos",      fill=fill)
            _cell(ws, row, 5, no_sets if first else "",  fill=fill, bold=first)
            _cell(ws, row, 6, f"{total_qty:.2f}", fill=fill)
            _cell(ws, row, 7, f"{unit_a:.2f}",    fill=fill)
            _cell(ws, row, 8, f"{row_area:.2f}",  fill=fill)
            first = False
            row  += 1

        # Total area row
        _cell(ws, row, 2, "", align="right")
        for ci in range(3, 7):
            ws.cell(row=row, column=ci).value = ""
        _cell(ws, row, 7, "Total Area (in SqM)", bold=True, align="right")
        _cell(ws, row, 8, f"{total_area:.2f}",   bold=True)
        row += 2   # blank gap

    # ── ACCESSORIES ───────────────────────────────────────────────────────────
    if hasattr(project, '_acc_agg') and project._acc_agg:
        acc_agg = project._acc_agg
    else:
        acc_agg = {}

    if acc_agg:
        ws.merge_cells(f"B{row}:H{row}")
        c = ws[f"B{row}"]
        c.value = "ACCESSORIES:"
        c.font  = Font(bold=True, size=9, color=_WHITE)
        c.fill  = PatternFill("solid", fgColor=_NAVY)
        c.alignment = Alignment(horizontal="left")
        ws.row_dimensions[row].height = 16
        row += 1

        for ci, h in enumerate(['PRODUCT', 'Qty', 'UOM'], start=2):
            _hdr(ws, row, ci, h, bg=_PURPLE)
        row += 1

        for sl, d in acc_agg.items():
            display, uom = _fmt_acc(sl)
            fill = _SMOKE if row % 2 == 0 else None
            _cell(ws, row, 2, display, align="left", fill=fill)
            _cell(ws, row, 3, f"{d['quantity']:.2f}", fill=fill)
            _cell(ws, row, 4, uom,  fill=fill)
            row += 1
        row += 1

    # ── SUMMARY ───────────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:H{row}")
    c = ws[f"B{row}"]
    c.value = "SUMMARY:"
    c.font  = Font(bold=True, size=9, color=_WHITE)
    c.fill  = PatternFill("solid", fgColor=_NAVY)
    c.alignment = Alignment(horizontal="left")
    ws.row_dimensions[row].height = 16
    row += 1

    sum_hdrs = ['PRODUCT', 'Total Quantity', 'UOM', 'Unit Area (SqM)', 'Total Area (SqM)']
    for ci, h in enumerate(sum_hdrs, start=2):
        _hdr(ws, row, ci, h, bg=_PURPLE)
    row += 1

    totals: dict[str, dict] = defaultdict(lambda: {'qty': 0, 'w': 0, 'h': 0})
    for boq in project.element_boqs:
        n = max(1, boq.element.quantity)
        for p in boq.panels:
            totals[p.size_label]['qty'] += p.quantity * n
            totals[p.size_label]['w']    = p.width_mm
            totals[p.size_label]['h']    = p.height_mm

    def _sort_key(kv):
        k = kv[0]
        return (0 if k.startswith('OC') or k.startswith('IC') else 1,
                -totals[k]['w'])

    grand_area = 0.0
    for k, d in sorted(totals.items(), key=_sort_key):
        unit_a = round(d['w'] * d['h'] / 1_000_000, 2)
        tot_a  = round(unit_a * d['qty'], 2)
        grand_area += tot_a
        fill = _SMOKE if row % 2 == 0 else None
        _cell(ws, row, 2, _fmt_panel(k),     align="left", fill=fill)
        _cell(ws, row, 3, f"{d['qty']:.2f}", fill=fill)
        _cell(ws, row, 4, "nos",             fill=fill)
        _cell(ws, row, 5, f"{unit_a:.2f}",   fill=fill)
        _cell(ws, row, 6, f"{tot_a:.2f}",    fill=fill)
        row += 1

    _cell(ws, row, 2, "", align="right")
    _cell(ws, row, 5, "Total Area", bold=True, align="right")
    _cell(ws, row, 6, f"{grand_area:.2f}", bold=True)

    # Column widths
    for col_letter, width in cw_map.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "B12"


# ══════════════════════════════════════════════════════════════════════════════
# QUOTATION sheet
# ══════════════════════════════════════════════════════════════════════════════

def _write_quotation_sheet(wb, project: ProjectBOQ, price_per_sqm: float = 0.0,
                           freight: float = 0.0, gst_rate: float = 0.18,
                           qtn_number: str = None, acc_agg: dict = None):
    ws  = wb.create_sheet("FORMWORK QUOTATION")
    row = 1

    today     = datetime.date.today()
    date_str  = project.date or today.strftime("%d/%m/%Y")
    valid_str = (today + datetime.timedelta(days=7)).strftime("%d/%m/%Y")
    qtn_num   = qtn_number or f"QTN-{abs(hash(project.project_name)) % 100000:05d}"

    # Title
    ws.merge_cells("B1:H1")
    c = ws["B1"]
    c.value = "FORMWORK QUOTATION"
    c.font  = Font(bold=True, size=14, color=_NIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for label, val in [("Quotation Number", qtn_num), ("Date", date_str), ("Valid Until", valid_str)]:
        ws.merge_cells(f"B{row+1}:H{row+1}")
        row += 1
        ws[f"B{row}"].value = f"{label}: {val}"
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].font = Font(size=9)

    row += 2   # gap

    # From / To
    from_info = [
        ("From:", True),
        ("NOVA FORMWORKS PVT LTD", True),
        ("A-7/121-124 South Side of GT Road Indl. Area, Ghaziabad (UP)", False),
        ("Email: hiren.jadav@novaformworks.com", False),
        ("Phone: 9211377073", False),
    ]
    to_info = [
        ("To:", True),
        (project.client_name or "—", True),
        (project.client_address or "—", False),
        (f"Email: {getattr(project, 'client_email', '') or '-'}", False),
        (f"Phone: {getattr(project, 'client_phone', '') or '-'}", False),
    ]
    for (fl, fb), (tl, tb) in zip(from_info, to_info):
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].value = fl
        ws[f"B{row}"].font  = Font(bold=fb, size=8)
        ws.merge_cells(f"F{row}:H{row}")
        ws[f"F{row}"].value = tl
        ws[f"F{row}"].font  = Font(bold=tb, size=8)
        ws[f"F{row}"].alignment = Alignment(horizontal="right")
        row += 1

    row += 1

    # QUOTATION DETAILS header
    ws.merge_cells(f"B{row}:H{row}")
    c = ws[f"B{row}"]
    c.value = "QUOTATION DETAILS:"
    c.font  = Font(bold=True, size=9, color=_WHITE)
    c.fill  = PatternFill("solid", fgColor=_NAVY)
    c.alignment = Alignment(horizontal="left")
    ws.row_dimensions[row].height = 16
    row += 1

    for ci, h in enumerate(['CLIENT REQUIREMENT', 'Total Area (in SqM)', 'Unit Price', 'Total Price'], start=2):
        _hdr(ws, row, ci, h, bg=_PURPLE)
    ws.row_dimensions[row].height = 22
    row += 1

    for boq in project.element_boqs:
        el   = boq.element
        n    = max(1, el.quantity)
        area = round(boq.total_panel_area_sqm * n, 2)
        price = round(area * price_per_sqm, 2) if price_per_sqm else 0.0
        fill = _SMOKE if row % 2 == 0 else None
        _cell(ws, row, 2, el.element_type.value.lower(), align="left", fill=fill)
        _cell(ws, row, 3, f"{area:.2f}",  fill=fill)
        _cell(ws, row, 4, f"{price_per_sqm:.2f}" if price_per_sqm else "0.00", fill=fill)
        _cell(ws, row, 5, f"{price:.2f}", fill=fill)
        row += 1

    row += 1

    # Accessories section
    if acc_agg:
        ws.merge_cells(f"B{row}:H{row}")
        ws[f"B{row}"].value = "2. Accessories Quotation"
        ws[f"B{row}"].font  = Font(bold=True, size=8, color=_WHITE)
        ws[f"B{row}"].fill  = PatternFill("solid", fgColor=_NAVY)
        ws.row_dimensions[row].height = 16
        row += 1

        for ci, h in enumerate(['Product', 'Quantity', 'UOM', 'Unit Price', 'Total Price'], start=2):
            _hdr(ws, row, ci, h, bg=_PURPLE)
        row += 1

        for sl, d in acc_agg.items():
            display, _ = _fmt_acc(sl)
            sl_up = sl.upper()
            if 'WALLER' in sl_up or 'TIE ROD' in sl_up or 'TIEROD' in sl_up:
                qty = d['total_length_m'] or d['quantity']
                uom = "mtr"
            else:
                qty = d['quantity']
                uom = "nos"
            fill = _SMOKE if row % 2 == 0 else None
            _cell(ws, row, 2, display, align="left", fill=fill)
            _cell(ws, row, 3, f"{qty:.2f}", fill=fill)
            _cell(ws, row, 4, uom,          fill=fill)
            _cell(ws, row, 5, "0.00",       fill=fill)
            _cell(ws, row, 6, "0.00",       fill=fill)
            row += 1
        row += 1

    # Financial summary
    from src.output.boq_generator import aggregate_project_boq
    agg = aggregate_project_boq(project)
    subtotal = round(agg['total_area_sqm'] * price_per_sqm, 2)
    gst_amt  = round(subtotal * gst_rate, 2)
    before_tax = round(subtotal + (freight or 0), 2)
    grand    = round(before_tax + gst_amt, 2)

    for label, val in [
        ("Subtotal",                 f"{subtotal:.2f}"),
        ("Freight",                  f"{freight or 0:.2f}"),
        ("Total Amount (Before Tax)",f"{before_tax:.2f}"),
        (f"GST ({gst_rate*100:.2f}%)",f"{gst_amt:.2f}"),
        ("Grand Total",              f"{grand:.2f}"),
    ]:
        ws.merge_cells(f"B{row}:F{row}")
        ws[f"B{row}"].value = label
        ws[f"B{row}"].alignment = Alignment(horizontal="right")
        is_grand = label == "Grand Total"
        ws[f"B{row}"].font = Font(bold=is_grand, size=8, color=_WHITE if is_grand else _NIGHT)
        if is_grand:
            ws[f"B{row}"].fill = PatternFill("solid", fgColor=_NAVY)
        c = ws.cell(row=row, column=7, value=val)
        c.font = Font(bold=is_grand, size=8)
        c.alignment = Alignment(horizontal="right")
        if is_grand:
            c.fill = PatternFill("solid", fgColor=_NAVY)
            c.font = Font(bold=True, size=8, color=_WHITE)
        row += 1

    row += 2

    # Terms & Conditions
    ws.merge_cells(f"B{row}:H{row}")
    ws[f"B{row}"].value = "TERMS AND CONDITIONS:"
    ws[f"B{row}"].font  = Font(bold=True, size=8, color=_NIGHT)
    ws.row_dimensions[row].height = 16
    row += 1

    tc = [
        "Freight Charges are in the client's scope.",
        "Accessories will be charged extra as per requirement.",
        "Payment: 50% advance with PO, balance 50% before dispatch.",
        "Delayed Payments: Simple interest @ 24% per annum.",
        "Dispatch within 7-10 days from drawing approval and advance receipt.",
        "Loading: Supplier's scope. Unloading: Buyer's scope.",
        "Training, travel, accommodation: Buyer's scope.",
        "Replacement: 1 new panel against 4 old panels.",
        "Taxes: GST and applicable government taxes charged extra.",
        "Jurisdiction: Delhi only.",
        "Validity: 7 days from date of issue.",
    ]
    for pt in tc:
        ws.merge_cells(f"B{row}:H{row}")
        ws[f"B{row}"].value = f"• {pt}"
        ws[f"B{row}"].font  = Font(size=7.5)
        ws[f"B{row}"].alignment = Alignment(horizontal="left", wrap_text=True)
        ws.row_dimensions[row].height = 14
        row += 1

    row += 1
    ws[f"B{row}"].value = "Prepared By:"
    ws[f"B{row}"].font  = Font(bold=True, size=8)
    row += 1
    ws[f"B{row}"].value = "Hiren Jadav"
    ws[f"B{row}"].font  = Font(size=8)

    # Column widths
    for col_letter, width in {'A': 4, 'B': 32, 'C': 14, 'D': 14, 'E': 14,
                               'F': 14, 'G': 14, 'H': 10}.items():
        ws.column_dimensions[col_letter].width = width


# ══════════════════════════════════════════════════════════════════════════════
# DAYS BOQ sheet (unchanged logic, updated colors)
# ══════════════════════════════════════════════════════════════════════════════

def _write_days_boq_sheet(wb, project: ProjectBOQ):
    ws = wb.create_sheet("DAYS BOQ")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "PANEL REUSE / DEPLOYMENT SCHEDULE"
    c.font  = Font(bold=True, size=13, color=_WHITE)
    c.fill  = PatternFill("solid", fgColor=_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:J2")
    ws["A2"].value = (
        f"Project: {project.project_name}   |   Client: {project.client_name}   |   "
        f"Generated: {datetime.date.today().strftime('%d-%m-%Y')}"
    )
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    panel_totals: dict[str, dict] = defaultdict(
        lambda: {"qty": 0, "w_mm": 0, "h_mm": 0, "elements": []})

    for boq in project.element_boqs:
        el = boq.element
        for panel in boq.panels:
            k = panel.size_label
            panel_totals[k]["qty"]      += panel.quantity * max(1, el.quantity)
            panel_totals[k]["w_mm"]      = panel.width_mm
            panel_totals[k]["h_mm"]      = panel.height_mm
            panel_totals[k]["elements"].append(el.label)

    def _sort_k(k):
        return (0 if k.startswith("OC") or k.startswith("IC") else 1,
                -panel_totals[k]["w_mm"])

    sorted_panels = sorted(panel_totals.keys(), key=_sort_k)

    row = 4
    headers = ["PANEL SIZE", "TOTAL INVENTORY", "DAY-1", "DAY-2", "DAY-3",
               "BALANCE", "ELEMENTS COVERED"]
    col_widths = [20, 18, 12, 12, 12, 12, 40]
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_PURPLE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 32
    row += 1

    _DAY_FILLS = ["DCE9F5", "E2EFDA", "FFF2CC"]
    for idx, key in enumerate(sorted_panels):
        d     = panel_totals[key]
        total = d["qty"]
        day1  = max(1, round(total / 3))
        day2  = max(1, round(total / 3))
        day3  = max(0, total - day1 - day2)
        bal   = total - day1 - day2 - day3
        elems = ", ".join(sorted(set(d["elements"]))[:10])
        is_corner = key.startswith("OC") or key.startswith("IC")
        bg = _LIGHT_P if is_corner else (_SMOKE if idx % 2 == 1 else None)

        ws.cell(row=row, column=1, value=_fmt_panel(key)).font = Font(bold=is_corner, size=8)
        if bg:
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=row, column=2, value=total).alignment = Alignment(horizontal="center")
        for ci, (dv, fc) in enumerate(zip([day1, day2, day3], _DAY_FILLS), start=3):
            c = ws.cell(row=row, column=ci, value=dv)
            c.fill = PatternFill("solid", fgColor=fc)
            c.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=bal).alignment  = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=elems).alignment = Alignment(horizontal="left", wrap_text=True)
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    note = ws[f"A{row}"]
    note.value = (
        "NOTE: Day-wise split is an equal 1/3 estimate. "
        "Adjust based on actual pour sequence and element priority on site."
    )
    note.font = Font(italic=True, size=9, color="595959")
    note.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel_boq(
    project: ProjectBOQ,
    output_path: str,
    price_per_sqm: float = 0.0,
    freight_amount: float = 0.0,
    gst_rate: float = 0.18,
    acc_agg: dict = None,
    boq_number: str = None,
    qtn_number: str = None,
) -> str:
    """
    Generate Excel file with three sheets:
    1. FORMWORK BOQ — panel quantities matching new template
    2. FORMWORK QUOTATION — pricing summary with T&C
    3. DAYS BOQ — panel deployment schedule
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Attach acc_agg to project temporarily for BOQ sheet
    project._acc_agg = acc_agg or {}

    _write_boq_sheet(wb, project, boq_number=boq_number)
    _write_quotation_sheet(wb, project, price_per_sqm=price_per_sqm,
                           freight=freight_amount, gst_rate=gst_rate,
                           qtn_number=qtn_number, acc_agg=acc_agg)
    _write_days_boq_sheet(wb, project)

    wb.save(output_path)
    return output_path
