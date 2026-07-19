"""
Panel Catalog Loader — dynamic panel width list with 3-tier priority:

  1. Central Excel  : {central_db_path}/panel_catalog.xlsx   (admin uploads, all users sync)
  2. Local Excel    : config/panel_catalog_local.xlsx         (admin sets on this machine)
  3. Default JSON   : config/panel_config.json                (shipped fallback)

Excel format (first sheet or sheet named "Panels"):
  Width_mm  |  Type          |  Note (optional)
  ----------|----------------|------------------
  600       |  FLAT          |
  500       |  FLAT          |
  ...       |  ...           |
  80        |  OC            |  Outer Corner
  100       |  IC            |  Inner Corner

Call reload() after uploading a new catalog to flush the in-memory cache.
"""

import json
import shutil
from pathlib import Path
from typing import Optional

_HERE         = Path(__file__).parent
_CONFIG_DIR   = _HERE.parent.parent / "config"
_CFG_JSON     = _CONFIG_DIR / "panel_config.json"
_LOCAL_XLSX   = _CONFIG_DIR / "panel_catalog_local.xlsx"
_TEMPLATE_XLSX = _CONFIG_DIR / "panel_catalog_template.xlsx"

CENTRAL_FILENAME = "panel_catalog.xlsx"

# In-memory cache: {flat_widths, oc_width, ic_width}
_cache: Optional[dict] = None
_cache_source: str = ""


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _load_from_excel(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    for name in wb.sheetnames:
        if name.strip().lower() in ("panels", "panel", "catalog"):
            ws = wb[name]
            break

    flat_widths: list[int] = []
    oc_width: Optional[int] = None
    ic_width: Optional[int] = None
    headers: list[str] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().upper().replace(" ", "_") if c else "" for c in row]
            continue
        if not any(row):
            continue
        rd = dict(zip(headers, row))
        raw_w = rd.get("WIDTH_MM") or rd.get("WIDTH")
        ptype  = str(rd.get("TYPE") or "FLAT").strip().upper()
        if raw_w is None:
            continue
        try:
            w = int(float(raw_w))
        except (ValueError, TypeError):
            continue
        if w <= 0:
            continue
        if ptype == "OC":
            oc_width = w
        elif ptype == "IC":
            ic_width = w
        else:
            flat_widths.append(w)

    wb.close()
    flat_widths = sorted(set(flat_widths), reverse=True)
    return {
        "flat_widths": flat_widths,
        "oc_width":    oc_width or 80,
        "ic_width":    ic_width or 100,
    }


def _load_from_json() -> dict:
    with open(_CFG_JSON) as f:
        cfg = json.load(f)
    ps = cfg["panel_system"]
    return {
        "flat_widths": sorted(ps["standard_widths_mm"], reverse=True),
        "oc_width":    ps.get("oc_width_mm", 80),
        "ic_width":    ps.get("ic_width_mm", 100),
    }


def _central_xlsx_path() -> Optional[Path]:
    try:
        api = _CONFIG_DIR / "api_config.json"
        if not api.exists():
            return None
        with open(api) as f:
            cfg = json.load(f)
        central = cfg.get("central_db_path", "").strip()
        if not central:
            return None
        p = Path(central) / CENTRAL_FILENAME
        return p if p.exists() else None
    except Exception:
        return None


# ── Cache management ───────────────────────────────────────────────────────────

def reload() -> None:
    global _cache, _cache_source
    _cache = None
    _cache_source = ""
    _ensure_loaded()


def _ensure_loaded() -> None:
    global _cache, _cache_source
    if _cache is not None:
        return

    # 1. Central Excel (admin-uploaded, shared across all users)
    central = _central_xlsx_path()
    if central:
        try:
            _cache = _load_from_excel(central)
            _cache_source = f"central:{central}"
            return
        except Exception:
            pass

    # 2. Local Excel override (admin-set on this machine)
    if _LOCAL_XLSX.exists():
        try:
            _cache = _load_from_excel(_LOCAL_XLSX)
            _cache_source = f"local:{_LOCAL_XLSX}"
            return
        except Exception:
            pass

    # 3. Fallback: panel_config.json
    _cache = _load_from_json()
    _cache_source = f"default:{_CFG_JSON}"


# ── Public accessors ───────────────────────────────────────────────────────────

def get_flat_widths() -> list:
    """Return flat panel widths sorted descending (e.g. [600,500,400,...])."""
    _ensure_loaded()
    return list(_cache["flat_widths"])


def get_usable_widths(min_width: int = 100) -> list:
    """Flat widths >= min_width (excludes spacers like 40mm)."""
    return [w for w in get_flat_widths() if w >= min_width]


def get_oc_width() -> int:
    _ensure_loaded()
    return _cache["oc_width"]


def get_ic_width() -> int:
    _ensure_loaded()
    return _cache["ic_width"]


def get_source() -> str:
    """Human-readable string describing where the active catalog was loaded from."""
    _ensure_loaded()
    src = _cache_source
    if src.startswith("central:"):
        return f"Central shared catalog  ({Path(src[8:]).name})"
    if src.startswith("local:"):
        return f"Local override  (panel_catalog_local.xlsx)"
    return "Default  (panel_config.json)"


def get_summary() -> str:
    """Short summary for UI display."""
    _ensure_loaded()
    fw = _cache["flat_widths"]
    max_w = fw[0] if fw else "—"
    return (
        f"Source : {get_source()}\n"
        f"Flat widths ({len(fw)}) : {', '.join(str(w) for w in fw)} mm\n"
        f"OC width : {_cache['oc_width']} mm   |   IC width : {_cache['ic_width']} mm\n"
        f"Widest panel : {max_w} mm"
    )


# ── Upload helpers (called from Admin Panel) ───────────────────────────────────

def upload_local_catalog(src_excel: str) -> tuple:
    """
    Validate and copy Excel to config/panel_catalog_local.xlsx.
    Returns (ok: bool, message: str).
    """
    try:
        data = _load_from_excel(Path(src_excel))
        if not data["flat_widths"]:
            return False, (
                "No FLAT panel widths found.\n"
                "Check column headers: Width_mm and Type must exist.\n"
                "Type values: FLAT, OC, IC"
            )
        shutil.copy2(src_excel, _LOCAL_XLSX)
        reload()
        fw = data["flat_widths"]
        return True, (
            f"Local catalog saved.\n"
            f"{len(fw)} flat widths: {', '.join(str(w) for w in fw)} mm\n"
            f"OC={data['oc_width']}mm  IC={data['ic_width']}mm"
        )
    except Exception as e:
        return False, f"Error reading Excel: {e}"


def upload_central_catalog(src_excel: str) -> tuple:
    """
    Validate and copy Excel to {{central_db_path}}/panel_catalog.xlsx.
    Returns (ok: bool, message: str).
    """
    try:
        api = _CONFIG_DIR / "api_config.json"
        with open(api) as f:
            cfg = json.load(f)
        central_str = cfg.get("central_db_path", "").strip()
        if not central_str:
            return False, (
                "Central DB path is not configured.\n"
                "Set it in Admin → Network tab first."
            )
        central_dir = Path(central_str)
        if not central_dir.exists():
            return False, f"Central path not accessible:\n{central_dir}"
        data = _load_from_excel(Path(src_excel))
        if not data["flat_widths"]:
            return False, "No FLAT panel widths found. Check column headers."
        dest = central_dir / CENTRAL_FILENAME
        shutil.copy2(src_excel, dest)
        reload()
        fw = data["flat_widths"]
        return True, (
            f"Central catalog updated.\nAll users will load this on next start.\n\n"
            f"{len(fw)} flat widths: {', '.join(str(w) for w in fw)} mm\n"
            f"OC={data['oc_width']}mm  IC={data['ic_width']}mm\n"
            f"Saved to: {dest}"
        )
    except Exception as e:
        return False, f"Error: {e}"


def clear_local_catalog() -> tuple:
    """Remove local override Excel. Returns (ok, message)."""
    try:
        if _LOCAL_XLSX.exists():
            _LOCAL_XLSX.unlink()
            reload()
            return True, "Local catalog override removed. Falling back to central or default."
        return True, "No local catalog was set."
    except Exception as e:
        return False, f"Could not remove local catalog: {e}"


def create_template_xlsx() -> Path:
    """
    Write a template Excel to config/panel_catalog_template.xlsx and return its path.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Panels"

    # Header row
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, hdr in enumerate(["Width_mm", "Type", "Note"], start=1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 35

    # Sample rows — current default widths
    rows = [
        (600, "FLAT", ""),
        (500, "FLAT", ""),
        (400, "FLAT", ""),
        (350, "FLAT", ""),
        (300, "FLAT", ""),
        (250, "FLAT", ""),
        (230, "FLAT", ""),
        (200, "FLAT", ""),
        (150, "FLAT", ""),
        (125, "FLAT", ""),
        (100, "FLAT", ""),
        (40,  "FLAT", "Spacer / fill plate"),
        (80,  "OC",   "Outer Corner — fixed 80mm"),
        (100, "IC",   "Inner Corner — fixed 100mm"),
    ]
    for r, (w, t, n) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=w)
        ws.cell(row=r, column=2, value=t)
        ws.cell(row=r, column=3, value=n)

    wb.save(str(_TEMPLATE_XLSX))
    return _TEMPLATE_XLSX
